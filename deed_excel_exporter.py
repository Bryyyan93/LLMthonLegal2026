"""
deed_excel_exporter.py

Generador de Excel para escrituras.
"""

from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font


EXCEL_COLUMNS = [
    ("id_fila", "INDICE"),
    ("tipo", "TIPO"),
    ("regimen", "RÉGIMEN"),
    ("descripcion", "DESCRIPCIÓN"),
    ("referencia_catastral", "REFERENCIA CATASTRAL"),
]


class ExcelExportError(Exception):
    """Error durante la generación del Excel."""
    pass


def flatten_deed(deed_result: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Convierte la estructura jerárquica de escritura
    en una lista plana de filas.
    """
    rows = []

    for item in deed_result.get("inventario", []):
        id_fila = item.get("id_fila")
        tipo = item.get("tipo")
        descripcion = item.get("descripcion")
        refs = item.get("referencias_catastrales", [])
        regimen = item.get("regimen")

        if not refs:
            rows.append({
                "id_fila": id_fila,
                "tipo": tipo,
                "regimen": regimen,
                "descripcion": descripcion,
                "referencia_catastral": None
            })
        else:
            for ref in refs:
                rows.append({
                    "id_fila": id_fila,
                    "tipo": tipo,
                    "regimen": regimen,
                    "descripcion": descripcion,
                    "referencia_catastral": ref
                })

    return rows


def export_deeds_to_excel(
    deed_result: Dict[str, Any],
    output_path: str
) -> None:

    rows = flatten_deed(deed_result)

    if not rows:
        raise ExcelExportError("No hay datos para exportar.")
    
    # Generar índice secuencial
    for idx, row in enumerate(rows, start=1):
        row["id_fila"] = idx

    wb = Workbook()
    ws = wb.active
    ws.title = "Escrituras"

    # Cabecera
    for col_index, (_, column_title) in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_index)
        cell.value = column_title
        cell.font = Font(bold=True)

    # Filas
    for row_index, row in enumerate(rows, start=2):
        for col_index, (field_key, _) in enumerate(EXCEL_COLUMNS, start=1):
            ws.cell(
                row=row_index,
                column=col_index,
                value=row.get(field_key)
            )

    # Ajuste ancho
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column].width = max_length + 2

    try:
        wb.save(output_path)
    except Exception as e:
        raise ExcelExportError(f"Error guardando Excel: {e}")
