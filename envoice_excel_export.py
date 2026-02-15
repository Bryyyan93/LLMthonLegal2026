"""
excel_exporter.py

Módulo encargado exclusivamente de generar el Excel final
a partir de una lista de facturas procesadas.
"""

from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font


EXCEL_COLUMNS = [
    ("numero_orden", "Nº ORDEN"),
    ("numero_factura", "Nº FACTURA"),
    ("fecha", "FECHA"),
    ("cif", "CIF"),
    ("proveedor", "PROVEEDOR"),
    ("comunidad_autonoma", "COMUNIDAD AUTÓNOMA"),
    ("articulo", "ARTÍCULO"),
    ("cantidad", "CANTIDAD"),
]


class ExcelExportError(Exception):
    """Error durante la generación del Excel."""
    pass


def export_invoices_to_excel(
    invoices: List[Dict[str, Any]],
    output_path: str
) -> None:
    """
    Genera un archivo Excel con las facturas procesadas.

    :param invoices: Lista de diccionarios validados.
    :param output_path: Ruta donde se guardará el .xlsx
    """

    if not invoices:
        raise ExcelExportError("No hay facturas para exportar.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # -----------------------------
    # Escribir cabecera
    # -----------------------------
    for col_index, (_, column_title) in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_index)
        cell.value = column_title
        cell.font = Font(bold=True)

    # -----------------------------
    # Escribir filas de datos
    # -----------------------------
    for row_index, invoice in enumerate(invoices, start=2):
        for col_index, (field_key, _) in enumerate(EXCEL_COLUMNS, start=1):
            ws.cell(
                row=row_index,
                column=col_index,
                value=invoice.get(field_key)
            )

    # Ajuste automático básico de ancho
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column].width = max_length + 2

    try:
        wb.save(output_path)
    except Exception as e:
        raise ExcelExportError(f"Error guardando Excel: {e}")
